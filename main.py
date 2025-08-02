import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
import os
from openpyxl.drawing.image import Image
import openpyxl
from openpyxl.utils import get_column_letter
import warnings

# Подавляем предупреждение о DataFrameGroupBy.apply
warnings.filterwarnings('ignore', message="DataFrameGroupBy.apply operated on the grouping columns")

# Создаем директории для данных и результатов, если они не существуют
os.makedirs('data', exist_ok=True)
os.makedirs('output', exist_ok=True)


def parse_month_value(value):
    """
    Преобразование значения месяца в дату
    Поддерживает только текстовый формат "Месяц Год"
    """
    if pd.isna(value):
        return None
    # Если значение строковое в формате "Месяц Год"
    if isinstance(value, str):
        try:
            # Словарь для преобразования русских названий месяцев
            month_dict = {
                'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6,
                'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12
            }
            parts = value.lower().split()
            if len(parts) == 2:
                month_name = parts[0]
                year = int(parts[1])
                month_num = month_dict.get(month_name)
                if month_num:
                    return datetime(year, month_num, 1)
        except:
            pass
    return None


def get_previous_month(month):
    """
    Возвращает предыдущий месяц в формате datetime
    """
    if month.month == 1:
        return datetime(month.year - 1, 12, 1)
    else:
        return datetime(month.year, month.month - 1, 1)


def get_projects_with_stop(financial_df, prolongations_df, month_columns):
    """
    Возвращает множество id проектов, у которых есть 'стоп' или 'end' в последнем месяце реализации или ранее
    """
    # Создаем список всех месяцев для проверки
    all_months = sorted(month_columns)

    # Преобразуем financial_df в длинный формат
    id_vars = [col for col in financial_df.columns if not isinstance(col, datetime)]
    long_df = financial_df.melt(id_vars=id_vars, var_name='financial_month', value_name='value')

    # Фильтруем только строки со значениями 'стоп' или 'end'
    stop_end_df = long_df[long_df['value'].isin(['стоп', 'end'])]

    # Объединяем с данными о пролонгациях
    merged_df = pd.merge(prolongations_df, stop_end_df, on='id', how='inner')

    # Фильтруем: месяц в financial_df должен быть <= или = последнему месяцу реализации
    merged_df = merged_df[merged_df['financial_month'] <= merged_df['month_date']]

    # Получаем уникальные id проектов
    projects_with_stop = set(merged_df['id'].unique())

    return projects_with_stop


def load_and_preprocess_data():
    """
    Загрузка и предобработка данных из файлов CSV
    """
    # Загрузка данных о пролонгациях
    prolongations_df = pd.read_csv('data/prolongations.csv')
    # Загрузка финансовых данных
    financial_df = pd.read_csv('data/financial_data.csv')

    # Предобработка данных о пролонгациях
    # Преобразование месяца в формат даты для удобства работы
    prolongations_df['month_date'] = prolongations_df['month'].apply(parse_month_value)
    # Удаляем строки, где не удалось преобразовать месяц
    prolongations_df = prolongations_df.dropna(subset=['month_date'])
    # Фильтруем только за 2023 год
    prolongations_df = prolongations_df[prolongations_df['month_date'].dt.year == 2023]
    # Исключаем записи с менеджером "без А/М"
    prolongations_df = prolongations_df[prolongations_df['AM'] != 'без А/М']

    # Предобработка финансовых данных
    # Определение колонок с месяцами
    month_columns = []
    for col in financial_df.columns:
        # Пропускаем служебные колонки
        if col in ['id', 'Причина дубля', 'Account']:
            continue
        # Проверяем, является ли колонка месяцем
        if parse_month_value(col) is not None:
            month_columns.append(col)

    # Преобразование названий колонок с месяцами в формат даты
    month_mapping = {}
    for col in month_columns:
        month_date = parse_month_value(col)
        if month_date:
            month_mapping[col] = month_date
    financial_df = financial_df.rename(columns=month_mapping)

    # Получение обновленного списка колонок с месяцами
    month_columns = [col for col in financial_df.columns if isinstance(col, datetime)]

    # Сортировка колонок с месяцами по хронологии
    month_columns_sorted = sorted(month_columns)

    # Переупорядочивание колонок в DataFrame
    other_columns = [col for col in financial_df.columns if col not in month_columns]
    financial_df = financial_df[other_columns + month_columns_sorted]

    # Агрегация данных для удаления дубликатов
    # Преобразуем в длинный формат
    id_vars = [col for col in financial_df.columns if not isinstance(col, datetime)]
    long_financial_df = financial_df.melt(id_vars=id_vars, var_name='financial_month', value_name='value')

    # Обработка значений: преобразуем числовые строки в числа, оставляем 'стоп', 'end', 'в ноль' как есть
    def clean_value(value):
        if pd.isna(value):
            return np.nan
        if value in ['стоп', 'end', 'в ноль']:
            return value
        try:
            # Преобразуем строку в число
            cleaned_value = str(value).replace('\xa0', '').replace(' ', '').replace(',', '.')
            return float(cleaned_value)
        except:
            return np.nan

    long_financial_df['value'] = long_financial_df['value'].apply(clean_value)

    # Агрегируем: для числовых значений суммируем, для текстовых берем первое значение
    def aggregate_values(group):
        numeric_values = group[pd.to_numeric(group['value'], errors='coerce').notna()]
        if not numeric_values.empty:
            return numeric_values['value'].sum()
        else:
            # Если нет числовых значений, берем первое не-NaN значение
            non_na_values = group[group['value'].notna()]
            if not non_na_values.empty:
                return non_na_values['value'].iloc[0]
            else:
                return np.nan

    # Используем group_keys=False для подавления предупреждения
    aggregated = long_financial_df.groupby(['id', 'financial_month'], group_keys=False).apply(
        aggregate_values).reset_index()
    aggregated.columns = ['id', 'financial_month', 'value']

    # Получаем не-месячные данные (без дубликатов по id)
    non_month_data = long_financial_df.groupby('id').first()
    non_month_data = non_month_data[[col for col in id_vars if col != 'id']].reset_index()

    # Преобразуем aggregated обратно в широкий формат
    wide_financial = aggregated.pivot_table(index='id', columns='financial_month', values='value',
                                            aggfunc='first').reset_index()

    # Объединяем с non_month_data
    financial_df = pd.merge(non_month_data, wide_financial, on='id', how='outer')

    # Восстанавливаем порядок столбцов
    other_columns = [col for col in non_month_data.columns if col != 'id']
    financial_df = financial_df[['id'] + other_columns + month_columns_sorted]

    return prolongations_df, financial_df, month_columns_sorted


def get_shipment_value_for_project(financial_df, project_id, month, for_denominator=False, debug=False, max_recursion=3,
                                   current_recursion=0):
    """
    Получение суммы отгрузки для конкретного проекта и месяца с учетом всех дублей
    for_denominator: если True, используем рекурсию для 'в ноль' (для знаменателя)
                    если False, не используем рекурсию (для числителя)
    """
    # Получаем все строки для проекта
    project_data = financial_df[financial_df['id'] == project_id]
    if month not in project_data.columns:
        if debug:
            print(f"  Проект {project_id}: месяц {month} отсутствует в данных")
        return 0

    # Получаем значение для указанного месяца
    value = project_data[month].iloc[0]
    if debug:
        print(f"  Проект {project_id}, месяц {month}: значение = {value}")

    # Проверяем значение
    if pd.isna(value) or value == '':
        if for_denominator and current_recursion < max_recursion:
            prev_month = get_previous_month(month)
            if prev_month is not None and prev_month in financial_df.columns:
                if debug:
                    print(
                        f"  Проект {project_id}: значение отсутствует, рекурсивный вызов для {prev_month} (знаменатель)")
                return get_shipment_value_for_project(
                    financial_df, project_id, prev_month, for_denominator, debug, max_recursion, current_recursion + 1
                )
            else:
                if debug:
                    print(f"  Проект {project_id}: предыдущий месяц {prev_month} отсутствует в данных")
                return 0
        else:
            if debug:
                print(f"  Проект {project_id}: значение отсутствует, возвращаем 0")
            return 0

    if value == 'в ноль':
        if for_denominator and current_recursion < max_recursion:
            prev_month = get_previous_month(month)
            if prev_month is not None and prev_month in financial_df.columns:
                if debug:
                    print(f"  Проект {project_id}: значение 'в ноль', рекурсивный вызов для {prev_month} (знаменатель)")
                return get_shipment_value_for_project(
                    financial_df, project_id, prev_month, for_denominator, debug, max_recursion, current_recursion + 1
                )
            else:
                if debug:
                    print(f"  Проект {project_id}: предыдущий месяц {prev_month} отсутствует в данных")
                return 0
        else:
            if debug:
                print(f"  Проект {project_id}: значение 'в ноль', возвращаем 0")
            return 0

    if value in ['стоп', 'end']:
        if debug:
            print(f"  Проект {project_id}: найдено значение '{value}'")
        return 0

    # Если значение числовое
    try:
        numeric_value = float(value)
        if debug:
            print(f"  Проект {project_id}: числовое значение {numeric_value}")
        return numeric_value
    except:
        if debug:
            print(f"  Проект {project_id}: не удалось преобразовать значение '{value}'")
        return 0


def get_shipment_sum(financial_df, projects_with_stop, project_ids, month, for_denominator=False, debug=False):
    """
    Получение суммы отгрузки для указанных проектов и месяца
    for_denominator: если True, используем рекурсию для 'в ноль' (для знаменателя)
                    если False, не используем рекурсию (для числителя)
    """
    if not project_ids:
        return 0
    # Проверяем, есть ли такой месяц в DataFrame
    if month not in financial_df.columns:
        return 0

    result = 0
    for project_id in project_ids:
        # Если проект исключен, пропускаем
        if project_id in projects_with_stop:
            if debug:
                print(f"  Проект {project_id} исключен из-за 'стоп' или 'end'")
            continue
        # Получаем сумму отгрузки для проекта и месяца
        shipment = get_shipment_value_for_project(financial_df, project_id, month, for_denominator, debug)
        result += shipment
        if debug:
            print(f"  Проект {project_id}: добавлено значение {shipment}, общая сумма = {result}")
    if debug:
        print(f"  Итоговая сумма для проектов {project_ids} в месяц {month}: {result}")
    return result


def calculate_prolongation_coefficients(prolongations_df, financial_df, month_columns):
    """
    Расчет коэффициентов пролонгации для каждого менеджера и для всего отдела
    """
    # Получаем множество проектов с 'стоп' или 'end'
    projects_with_stop = get_projects_with_stop(financial_df, prolongations_df, month_columns)
    # Выводим отладочную информацию
    print(f"Проекты с 'стоп' или 'end' в последнем месяце реализации или ранее: {len(projects_with_stop)}")
    print(f"Первые 5 проектов с 'стоп' или 'end': {list(projects_with_stop)[:5]}")

    # Получаем уникальных менеджеров
    managers = prolongations_df['AM'].unique()
    # Получаем уникальные месяцы из данных о пролонгациях
    unique_months = sorted(prolongations_df['month_date'].unique())

    # Создаем DataFrame для хранения результатов
    results = []

    # Рассчитываем коэффициенты для каждого месяца
    for i in range(1, len(unique_months)):
        current_month = unique_months[i]
        previous_month = unique_months[i - 1]

        # Проверяем, есть ли текущий и предыдущий месяцы в финансовых данных
        if current_month not in financial_df.columns or previous_month not in financial_df.columns:
            continue

        # Для расчета коэффициента за первый месяц нам нужны проекты, завершившиеся в предыдущем месяце
        prev_month_projects = prolongations_df[prolongations_df['month_date'] == previous_month]

        # Для расчета коэффициента за второй месяц нам нужны проекты, завершившиеся два месяца назад
        if i > 1:
            two_months_ago = unique_months[i - 2]
            two_months_ago_projects = prolongations_df[prolongations_df['month_date'] == two_months_ago]
        else:
            two_months_ago_projects = pd.DataFrame()

        # Рассчитываем коэффициенты для каждого менеджера
        for manager in managers:
            # Проекты менеджера, завершившиеся в предыдущем месяце
            manager_prev_month_projects = prev_month_projects[prev_month_projects['AM'] == manager]

            # Проекты менеджера, завершившиеся два месяца назад
            if not two_months_ago_projects.empty:
                manager_two_months_ago_projects = two_months_ago_projects[two_months_ago_projects['AM'] == manager]
            else:
                manager_two_months_ago_projects = pd.DataFrame()

            # Расчет коэффициента для первого месяца
            if not manager_prev_month_projects.empty:
                # Сумма отгрузки за последний месяц реализации проектов (знаменатель)
                ids = manager_prev_month_projects['id'].tolist()
                # Включаем отладку для первых 3 проектов каждого менеджера
                debug = len(results) < 10  # Отладка для первых 10 записей
                if debug:
                    print(f"\nОтладка для менеджера {manager}, месяц {current_month}:")
                    print(f"Проекты, завершившиеся в {previous_month}: {ids[:3]}...")

                # Для знаменателя используем рекурсию для 'в ноль'
                last_month_shipment = get_shipment_sum(
                    financial_df, projects_with_stop, ids, previous_month, for_denominator=True, debug=debug
                )

                # Для числителя не используем рекурсию для 'в ноль'
                current_month_shipment = get_shipment_sum(
                    financial_df, projects_with_stop, ids, current_month, for_denominator=False, debug=debug
                )

                # Коэффициент пролонгации для первого месяца
                if last_month_shipment > 0:
                    first_month_coeff = current_month_shipment / last_month_shipment
                else:
                    first_month_coeff = 0

                if debug:
                    print(f"Результат для менеджера {manager}:")
                    print(f"  Сумма за {previous_month} (знаменатель): {last_month_shipment}")
                    print(f"  Сумма за {current_month} (числитель): {current_month_shipment}")
                    print(f"  Коэффициент: {first_month_coeff}")
            else:
                last_month_shipment = 0
                current_month_shipment = 0
                first_month_coeff = 0

            # Расчет коэффициента для второго месяца
            if not manager_two_months_ago_projects.empty:
                # Проверяем, есть ли месяц два месяца назад в финансовых данных
                if two_months_ago in financial_df.columns:
                    # Получаем ID проектов, завершившихся два месяца назад
                    ids = manager_two_months_ago_projects['id'].tolist()

                    # Отфильтровываем проекты, у которых не было отгрузки в предыдущем месяце
                    not_prolonged_first_month_ids = []
                    for proj_id in ids:
                        prev_month_shipment = get_shipment_sum(
                            financial_df, projects_with_stop, [proj_id], previous_month, for_denominator=False
                        )
                        if prev_month_shipment == 0:
                            not_prolonged_first_month_ids.append(proj_id)

                    # Сумма отгрузки за последний месяц реализации проектов, не пролонгированных в первый месяц (знаменатель)
                    last_month_shipment_second = get_shipment_sum(
                        financial_df, projects_with_stop, not_prolonged_first_month_ids,
                        two_months_ago, for_denominator=True
                    )

                    # Сумма отгрузки за текущий месяц (пролонгация во второй месяц) (числитель)
                    current_month_shipment_second = get_shipment_sum(
                        financial_df, projects_with_stop, not_prolonged_first_month_ids,
                        current_month, for_denominator=False
                    )

                    # Коэффициент пролонгации для второго месяца
                    if last_month_shipment_second > 0:
                        second_month_coeff = current_month_shipment_second / last_month_shipment_second
                    else:
                        second_month_coeff = 0
                else:
                    last_month_shipment_second = 0
                    current_month_shipment_second = 0
                    second_month_coeff = 0
            else:
                last_month_shipment_second = 0
                current_month_shipment_second = 0
                second_month_coeff = 0

            # Добавляем результаты для текущего менеджера и месяца
            results.append({
                'Manager': manager,
                'Month': current_month,
                'First_Month_Coeff': first_month_coeff,
                'Second_Month_Coeff': second_month_coeff,
                'First_Month_Shipment': current_month_shipment,
                'Second_Month_Shipment': current_month_shipment_second,
                'Prev_Month_Shipment': last_month_shipment,
                'Two_Months_Ago_Shipment': last_month_shipment_second
            })

    # Создаем DataFrame с результатами
    results_df = pd.DataFrame(results)

    # Рассчитываем общие коэффициенты по отделу
    dept_results = []
    for i in range(1, len(unique_months)):
        current_month = unique_months[i]
        previous_month = unique_months[i - 1]

        # Проверяем, есть ли текущий и предыдущий месяцы в финансовых данных
        if current_month not in financial_df.columns or previous_month not in financial_df.columns:
            continue

        # Для расчета коэффициента за первый месяц нам нужны проекты, завершившиеся в предыдущем месяце
        prev_month_projects = prolongations_df[prolongations_df['month_date'] == previous_month]

        # Для расчета коэффициента за второй месяц нам нужны проекты, завершившиеся два месяца назад
        if i > 1:
            two_months_ago = unique_months[i - 2]
            two_months_ago_projects = prolongations_df[prolongations_df['month_date'] == two_months_ago]
        else:
            two_months_ago_projects = pd.DataFrame()

        # Расчет коэффициента для первого месяца
        if not prev_month_projects.empty:
            # Сумма отгрузки за последний месяц реализации проектов (знаменатель)
            ids = prev_month_projects['id'].tolist()

            # Для знаменателя используем рекурсию для 'в ноль'
            last_month_shipment = get_shipment_sum(
                financial_df, projects_with_stop, ids, previous_month, for_denominator=True
            )

            # Для числителя не используем рекурсию для 'в ноль'
            current_month_shipment = get_shipment_sum(
                financial_df, projects_with_stop, ids, current_month, for_denominator=False
            )

            # Коэффициент пролонгации для первого месяца
            if last_month_shipment > 0:
                first_month_coeff = current_month_shipment / last_month_shipment
            else:
                first_month_coeff = 0
        else:
            last_month_shipment = 0
            current_month_shipment = 0
            first_month_coeff = 0

        # Расчет коэффициента для второго месяца
        if not two_months_ago_projects.empty:
            # Проверяем, есть ли месяц два месяца назад в финансовых данных
            if two_months_ago in financial_df.columns:
                # Получаем ID проектов, завершившихся два месяца назад
                ids = two_months_ago_projects['id'].tolist()

                # Отфильтровываем проекты, у которых не было отгрузки в предыдущем месяце
                not_prolonged_first_month_ids = []
                for proj_id in ids:
                    prev_month_shipment = get_shipment_sum(
                        financial_df, projects_with_stop, [proj_id], previous_month, for_denominator=False
                    )
                    if prev_month_shipment == 0:
                        not_prolonged_first_month_ids.append(proj_id)

                # Сумма отгрузки за последний месяц реализации проектов, не пролонгированных в первый месяц (знаменатель)
                last_month_shipment_second = get_shipment_sum(
                    financial_df, projects_with_stop, not_prolonged_first_month_ids,
                    two_months_ago, for_denominator=True
                )

                # Сумма отгрузки за текущий месяц (пролонгация во второй месяц) (числитель)
                current_month_shipment_second = get_shipment_sum(
                    financial_df, projects_with_stop, not_prolonged_first_month_ids,
                    current_month, for_denominator=False
                )

                # Коэффициент пролонгации для второго месяца
                if last_month_shipment_second > 0:
                    second_month_coeff = current_month_shipment_second / last_month_shipment_second
                else:
                    second_month_coeff = 0
            else:
                last_month_shipment_second = 0
                current_month_shipment_second = 0
                second_month_coeff = 0
        else:
            last_month_shipment_second = 0
            current_month_shipment_second = 0
            second_month_coeff = 0

        # Добавляем результаты для отдела и текущего месяца
        dept_results.append({
            'Manager': 'Отдел',
            'Month': current_month,
            'First_Month_Coeff': first_month_coeff,
            'Second_Month_Coeff': second_month_coeff,
            'First_Month_Shipment': current_month_shipment,
            'Second_Month_Shipment': current_month_shipment_second,
            'Prev_Month_Shipment': last_month_shipment,
            'Two_Months_Ago_Shipment': last_month_shipment_second
        })

    # Создаем DataFrame с результатами по отделу
    dept_results_df = pd.DataFrame(dept_results)

    # Объединяем результаты
    all_results_df = pd.concat([results_df, dept_results_df], ignore_index=True)

    # Рассчитываем годовые коэффициенты
    yearly_results = calculate_yearly_coefficients(all_results_df)

    return all_results_df, yearly_results


def calculate_yearly_coefficients(monthly_results_df):
    """
    Расчет годовых коэффициентов пролонгации на основе месячных данных
    """
    # Группируем данные по менеджерам
    grouped = monthly_results_df.groupby('Manager')
    yearly_results = []

    for manager, group in grouped:
        # Пропускаем отдел, так как он будет рассчитан отдельно
        if manager == 'Отдел':
            continue

        # Суммируем значения за год
        total_first_shipment = group['First_Month_Shipment'].sum()
        total_second_shipment = group['Second_Month_Shipment'].sum()
        total_prev_month_shipment = group['Prev_Month_Shipment'].sum()
        total_two_months_ago_shipment = group['Two_Months_Ago_Shipment'].sum()

        # Рассчитываем годовые коэффициенты
        if total_prev_month_shipment > 0:
            first_month_coeff = total_first_shipment / total_prev_month_shipment
        else:
            first_month_coeff = 0

        if total_two_months_ago_shipment > 0:
            second_month_coeff = total_second_shipment / total_two_months_ago_shipment
        else:
            second_month_coeff = 0

        yearly_results.append({
            'Manager': manager,
            'First_Month_Coeff': first_month_coeff,
            'Second_Month_Coeff': second_month_coeff,
            'First_Month_Shipment': total_first_shipment,
            'Second_Month_Shipment': total_second_shipment,
            'Prev_Month_Shipment': total_prev_month_shipment,
            'Two_Months_Ago_Shipment': total_two_months_ago_shipment
        })

    # Рассчитываем годовые коэффициенты для отдела
    dept_data = monthly_results_df[monthly_results_df['Manager'] == 'Отдел']
    total_first_shipment = dept_data['First_Month_Shipment'].sum()
    total_second_shipment = dept_data['Second_Month_Shipment'].sum()
    total_prev_month_shipment = dept_data['Prev_Month_Shipment'].sum()
    total_two_months_ago_shipment = dept_data['Two_Months_Ago_Shipment'].sum()

    if total_prev_month_shipment > 0:
        first_month_coeff = total_first_shipment / total_prev_month_shipment
    else:
        first_month_coeff = 0

    if total_two_months_ago_shipment > 0:
        second_month_coeff = total_second_shipment / total_two_months_ago_shipment
    else:
        second_month_coeff = 0

    yearly_results.append({
        'Manager': 'Отдел',
        'First_Month_Coeff': first_month_coeff,
        'Second_Month_Coeff': second_month_coeff,
        'First_Month_Shipment': total_first_shipment,
        'Second_Month_Shipment': total_second_shipment,
        'Prev_Month_Shipment': total_prev_month_shipment,
        'Two_Months_Ago_Shipment': total_two_months_ago_shipment
    })

    # Создаем DataFrame с годовыми результатами
    yearly_results_df = pd.DataFrame(yearly_results)

    return yearly_results_df


def auto_adjust_columns(worksheet):
    """Автоподбор ширины столбцов"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_analytical_report(monthly_results_df, yearly_results_df, prolongations_df, financial_df, projects_with_stop):
    """
    Создание аналитического отчета в Excel
    """
    # Создаем Excel writer
    writer = pd.ExcelWriter('output/prolongation_report.xlsx', engine='openpyxl')

    # Преобразуем даты в строковый формат для удобства чтения
    monthly_report = monthly_results_df.copy()
    monthly_report['Month'] = monthly_report['Month'].dt.strftime('%Y-%m')

    # Округляем значения
    monthly_report['First_Month_Coeff'] = monthly_report['First_Month_Coeff'].round(3)
    monthly_report['Second_Month_Coeff'] = monthly_report['Second_Month_Coeff'].round(3)
    monthly_report['First_Month_Shipment'] = monthly_report['First_Month_Shipment'].round(0)
    monthly_report['Second_Month_Shipment'] = monthly_report['Second_Month_Shipment'].round(0)
    monthly_report['Prev_Month_Shipment'] = monthly_report['Prev_Month_Shipment'].round(0)
    monthly_report['Two_Months_Ago_Shipment'] = monthly_report['Two_Months_Ago_Shipment'].round(0)

    # Записываем месячные коэффициенты
    monthly_report.to_excel(writer, sheet_name='Месячные коэффициенты', index=False)

    # Округляем годовые значения
    yearly_report = yearly_results_df.copy()
    yearly_report['First_Month_Coeff'] = yearly_report['First_Month_Coeff'].round(3)
    yearly_report['Second_Month_Coeff'] = yearly_report['Second_Month_Coeff'].round(3)
    yearly_report['First_Month_Shipment'] = yearly_report['First_Month_Shipment'].round(0)
    yearly_report['Second_Month_Shipment'] = yearly_report['Second_Month_Shipment'].round(0)
    yearly_report['Prev_Month_Shipment'] = yearly_report['Prev_Month_Shipment'].round(0)
    yearly_report['Two_Months_Ago_Shipment'] = yearly_report['Two_Months_Ago_Shipment'].round(0)

    # Записываем годовые коэффициенты
    yearly_report.to_excel(writer, sheet_name='Годовые коэффициенты', index=False)

    # Создаем сводную таблицу по месяцам и менеджерам
    pivot_first = monthly_report.pivot_table(
        values='First_Month_Coeff',
        index='Manager',
        columns='Month'
    )

    pivot_second = monthly_report.pivot_table(
        values='Second_Month_Coeff',
        index='Manager',
        columns='Month'
    )

    # Записываем сводные таблицы
    pivot_first.to_excel(writer, sheet_name='Сводная по первому месяцу')
    pivot_second.to_excel(writer, sheet_name='Сводная по второму месяцу')

    # Количество проектов каждого менеджера
    project_counts = prolongations_df.groupby('AM')['id'].nunique().reset_index()
    project_counts.columns = ['Manager', 'Количество проектов']
    project_counts.to_excel(writer, sheet_name='Количество проектов', index=False)

    # Средние коэффициенты по менеджерам (взвешенное среднее)
    managers_data = monthly_results_df[monthly_results_df['Manager'] != 'Отдел']
    unique_managers = managers_data['Manager'].unique()
    avg_coeffs_list = []

    for manager in unique_managers:
        manager_data = managers_data[managers_data['Manager'] == manager]

        # Расчет взвешенного среднего для первого месяца
        if manager_data['Prev_Month_Shipment'].sum() > 0:
            first_month_avg = np.average(
                manager_data['First_Month_Coeff'],
                weights=manager_data['Prev_Month_Shipment']
            )
        else:
            first_month_avg = 0

        # Расчет взвешенного среднего для второго месяца
        if manager_data['Two_Months_Ago_Shipment'].sum() > 0:
            second_month_avg = np.average(
                manager_data['Second_Month_Coeff'],
                weights=manager_data['Two_Months_Ago_Shipment']
            )
        else:
            second_month_avg = 0

        avg_coeffs_list.append({
            'Manager': manager,
            'First_Month_Coeff': round(first_month_avg, 3),
            'Second_Month_Coeff': round(second_month_avg, 3)
        })

    avg_coeffs = pd.DataFrame(avg_coeffs_list)
    avg_coeffs.to_excel(writer, sheet_name='Средние месячные коэффициенты', index=False)

    # Создаем лист с исключенными проектами
    excluded_projects = []
    for project_id in projects_with_stop:
        # Находим информацию о проекте
        project_info = prolongations_df[prolongations_df['id'] == project_id]
        if not project_info.empty:
            manager = project_info['AM'].iloc[0]
            last_month = project_info['month_date'].iloc[0]

            # Находим причину исключения
            project_financial = financial_df[financial_df['id'] == project_id]
            reason = None

            # Проверяем каждый месяц до последнего месяца реализации
            for month in sorted([col for col in project_financial.columns if isinstance(col, datetime)]):
                if month > last_month:
                    continue

                for _, row in project_financial.iterrows():
                    value = row[month]
                    if value in ['стоп', 'end']:
                        reason = value
                        break
                if reason:
                    break

            excluded_projects.append({
                'Project ID': project_id,
                'Manager': manager,
                'Last Month': last_month.strftime('%Y-%m'),
                'Reason': reason
            })

    excluded_df = pd.DataFrame(excluded_projects)
    excluded_df.to_excel(writer, sheet_name='Исключенные проекты', index=False)

    # Создаем визуализации
    create_visualizations(monthly_report, yearly_results_df, writer)

    # Добавляем лист с пояснениями
    workbook = writer.book
    explanation_sheet = workbook.create_sheet('Пояснения')
    explanations = [
        ["Пояснения к отчету о пролонгациях"],
        [""],
        ["Коэффициент пролонгации для первого месяца:"],
        ["Отношение суммы отгрузки проектов, пролонгированных в первый месяц после завершения,"],
        ["к сумме отгрузки последнего месяца реализации всех завершившихся в прошлом месяце проектов."],
        [""],
        ["Коэффициент пролонгации для второго месяца:"],
        ["Отношение суммы отгрузки проектов, пролонгированных во второй месяц,"],
        ["к сумме отгрузки последнего месяца проектов, не пролонгированных в первый месяц."],
        [""],
        ["Особенности обработки данных:"],
        ["- Проекты с отметками 'стоп' или 'end' исключаются из расчета"],
        ["- Если в последнем месяце реализации все значения 'в ноль', для знаменателя"],
        ["  берется отгрузка предыдущего месяца (если доступна)"],
        ["- Для числителя (пролонгация) отгрузка берется только за указанный месяц"],
        ["- Отчет включает данные только за 2023 год"],
        ["- Записи с менеджером 'без А/М' исключены из расчета"],
        [""],
        ["Пояснения к листам отчета:"],
        ["- 'Месячные коэффициенты' - помесячные коэффициенты по каждому менеджеру и отделу"],
        ["- 'Годовые коэффициенты' - итоговые коэффициенты за весь 2023 год"],
        ["- 'Средние месячные коэффициенты' - взвешенные средние помесячные коэффициенты"],
        ["- 'Количество проектов' - количество уникальных проектов у каждого менеджера"],
        ["- 'Исключенные проекты' - проекты, исключенные из расчета из-за 'стоп' или 'end'"]
    ]

    for row in explanations:
        explanation_sheet.append(row)

    # Сохраняем Excel файл
    writer.close()

    # Открываем файл для автоподбора ширины столбцов
    workbook = openpyxl.load_workbook('output/prolongation_report.xlsx')

    # Автоподбор ширины столбцов для каждого листа
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        auto_adjust_columns(worksheet)

    # Сохраняем изменения
    workbook.save('output/prolongation_report.xlsx')
    print("Аналитический отчет сохранен в файле output/prolongation_report.xlsx")


def create_visualizations(monthly_report, yearly_results_df, writer):
    """
    Создание визуализаций для отчета
    """
    # Используем доступный стиль matplotlib
    plt.style.use('ggplot')

    # Фильтруем данные, исключая отдел
    managers_data = monthly_report[monthly_report['Manager'] != 'Отдел']

    # Создаем фигуру с несколькими подграфиками
    fig = plt.figure(figsize=(15, 10))

    # График 1: Коэффициенты пролонгации по месяцам для каждого менеджера (первый месяц)
    ax1 = fig.add_subplot(221)
    for manager in managers_data['Manager'].unique():
        manager_data = managers_data[managers_data['Manager'] == manager]
        ax1.plot(manager_data['Month'], manager_data['First_Month_Coeff'], marker='o', label=manager)
    ax1.set_title('Коэффициенты пролонгации (первый месяц)', fontsize=12)
    ax1.set_xlabel('Месяц', fontsize=10)
    ax1.set_ylabel('Коэффициент', fontsize=10)
    ax1.legend()
    ax1.grid(True)
    plt.xticks(rotation=45)

    # График 2: Коэффициенты пролонгации по месяцам для каждого менеджера (второй месяц)
    ax2 = fig.add_subplot(222)
    for manager in managers_data['Manager'].unique():
        manager_data = managers_data[managers_data['Manager'] == manager]
        ax2.plot(manager_data['Month'], manager_data['Second_Month_Coeff'], marker='o', label=manager)
    ax2.set_title('Коэффициенты пролонгации (второй месяц)', fontsize=12)
    ax2.set_xlabel('Месяц', fontsize=10)
    ax2.set_ylabel('Коэффициент', fontsize=10)
    ax2.legend()
    ax2.grid(True)
    plt.xticks(rotation=45)

    # График 3: Годовые коэффициенты по менеджерам
    ax3 = fig.add_subplot(223)
    managers_yearly = yearly_results_df[yearly_results_df['Manager'] != 'Отдел']
    x = np.arange(len(managers_yearly['Manager']))
    width = 0.35
    ax3.bar(x - width / 2, managers_yearly['First_Month_Coeff'], width, label='Первый месяц')
    ax3.bar(x + width / 2, managers_yearly['Second_Month_Coeff'], width, label='Второй месяц')
    ax3.set_title('Годовые коэффициенты пролонгации по менеджерам', fontsize=12)
    ax3.set_xlabel('Менеджер', fontsize=10)
    ax3.set_ylabel('Коэффициент', fontsize=10)
    ax3.set_xticks(x)
    ax3.set_xticklabels(managers_yearly['Manager'], rotation=45)
    ax3.legend()
    ax3.grid(True)

    # График 4: Сравнение менеджеров по средним коэффициентам
    ax4 = fig.add_subplot(224)
    avg_coeffs = managers_data.groupby('Manager').agg({
        'First_Month_Coeff': 'mean',
        'Second_Month_Coeff': 'mean'
    }).reset_index()
    x = np.arange(len(avg_coeffs['Manager']))
    width = 0.35
    ax4.bar(x - width / 2, avg_coeffs['First_Month_Coeff'], width, label='Первый месяц')
    ax4.bar(x + width / 2, avg_coeffs['Second_Month_Coeff'], width, label='Второй месяц')
    ax4.set_title('Средние коэффициенты пролонгации по менеджерам', fontsize=12)
    ax4.set_xlabel('Менеджер', fontsize=10)
    ax4.set_ylabel('Коэффициент', fontsize=10)
    ax4.set_xticks(x)
    ax4.set_xticklabels(avg_coeffs['Manager'], rotation=45)
    ax4.legend()
    ax4.grid(True)

    plt.tight_layout()

    # Сохраняем график
    plt.savefig('output/prolongation_visualizations.png')

    # Добавляем график в Excel
    workbook = writer.book
    worksheet = workbook.create_sheet('Визуализации')
    img = Image('output/prolongation_visualizations.png')
    worksheet.add_image(img, 'A1')

    plt.close()


def main():
    """
    Основная функция для выполнения анализа
    """
    print("Начало анализа коэффициентов пролонгации...")

    # Загрузка и предобработка данных
    prolongations_df, financial_df, month_columns = load_and_preprocess_data()

    # Выводим отладочную информацию
    print(f"Уникальные месяцы в данных о пролонгациях: {prolongations_df['month_date'].unique()}")

    # Получаем месяцы из финансовых данных
    financial_months = [col for col in financial_df.columns if isinstance(col, datetime)]
    print(f"Месяцы в финансовых данных: {financial_months}")

    # Расчет коэффициентов пролонгации
    monthly_results_df, yearly_results_df = calculate_prolongation_coefficients(
        prolongations_df, financial_df, month_columns
    )

    # Получаем множество проектов с 'стоп' или 'end'
    projects_with_stop = get_projects_with_stop(financial_df, prolongations_df, month_columns)

    # Создание аналитического отчета
    create_analytical_report(monthly_results_df, yearly_results_df, prolongations_df, financial_df, projects_with_stop)

    print("Анализ завершен!")


if __name__ == "__main__":
    main()